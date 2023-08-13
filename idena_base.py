import json
import pro_base
import openpyxl
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.cvm.v20170312 import cvm_client, models
state_excel = 'idena-all.xlsx'
def readfile(filename):
    with open(filename, 'r') as f:
        content = f.read()  # 读取文件内容
    return content  # 返回文件内容

secretId = readfile('secretId')  # 读取 example.txt 中的内容并赋值给 secretId
secretKey = readfile('secretKey')  # 读取 example.txt 中的内容并赋值给 secretKey


def img_vpsnamesearch():
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.DescribeInstancesRequest()
        params = {
            "Limit": 100,
            "Filters": [
                {
                    "Name": "tag-key",
                    "Values": ["idena"]
                },
                {
                    "Name": "tag-value",
                    "Values": ["img"]
                }
            ]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个DescribeInstancesResponse的实例，与请求对象对应
        resp = client.DescribeInstances(req)

        a = json.loads(resp.to_json_string())
        b = a['InstanceSet']
        c = int(len(b) - 1)
        while c >= 0:
            print(b[c]['InstanceId'], b[c]['CreatedTime'], b[c]['PublicIpAddresses'])
            c -= 1
        print()

    except TencentCloudSDKException as err:
        print(err)

def vps_build():
    Password = input("请输入密码，直接按enter默认Re123456789123") or "Re123456789123"
    print("-" * 5)
    times = int(input("请输入服务器个数"))
    print("-" * 5)
    print("idena镜像列表")
    img_namesearch()
    img = input("请输入镜像")
    print("-" * 5)
    print("自动今晚11：40注销服务器")
    print("请牢记密码为", Password)
    print("-" * 5)
    while times > 0:
        times -= 1
        try:
            # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
            # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
            # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
            cred = credential.Credential(secretId, secretKey)
            # 实例化一个http选项，可选的，没有特殊需求可以跳过
            httpProfile = HttpProfile()
            httpProfile.endpoint = "cvm.ap-guangzhou.tencentcloudapi.com"

            # 实例化一个client选项，可选的，没有特殊需求可以跳过
            clientProfile = ClientProfile()
            clientProfile.httpProfile = httpProfile
            # 实例化要请求产品的client对象,clientProfile是可选的
            client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

            # 实例化一个请求对象,每个接口都会对应一个request对象
            req = models.RunInstancesRequest()
            params = {
                "InstanceChargeType": "POSTPAID_BY_HOUR",
                "DisableApiTermination": False,
                "Placement": {
                    "Zone": "ap-guangzhou-6",
                    "ProjectId": 0
                },
                "VirtualPrivateCloud": {
                    "AsVpcGateway": False,
                    "VpcId": "vpc-b8ptxop5",
                    "SubnetId": "subnet-ahm3l13o",
                    "Ipv6AddressCount": 0
                },
                "InstanceType": "S6.MEDIUM8",
                "ImageId": img,
                "SystemDisk": {
                    "DiskSize": 50,
                    "DiskType": "CLOUD_BSSD"
                },
                "InternetAccessible": {
                    "InternetMaxBandwidthOut": 100,
                    "PublicIpAssigned": True,
                    "InternetChargeType": "TRAFFIC_POSTPAID_BY_HOUR"
                },
                "LoginSettings": {
                    "Password": Password
                },
                "SecurityGroupIds": ["sg-dbcg6j0n"],
                "InstanceCount": 1,
                "EnhancedService": {
                    "SecurityService": {
                        "Enabled": False
                    },
                    "MonitorService": {
                        "Enabled": False
                    },
                    "AutomationService": {
                        "Enabled": True
                    }
                },
                "ActionTimer": {
                    "ActionTime": pro_base.today_day() + " 23:40:00",
                    "TimerAction": "TerminateInstances",
                    "Externals": {
                        "ReleaseAddress": True
                    }
                },
                "TagSpecification": [
                    {
                        "ResourceType": "instance",
                        "Tags": [
                            {
                                "Key": "idena",
                                "Value": "idena"
                            }
                        ]
                    }
                ]
            }
            req.from_json_string(json.dumps(params))

            # 返回的resp是一个RunInstancesResponse的实例，与请求对象对应
            resp = client.RunInstances(req)
            # 输出json格式的字符串回包
            print(resp.to_json_string())

        except TencentCloudSDKException as err:
            print(err)
#定义函数
def img_namesearch():
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.DescribeImagesRequest()
        params = {
             "Filters": [
            {
                "Name": "tag-key",
                "Values": [ "idena" ]
            },
            {
                "Name": "tag-value",
                "Values": [ "idena" ]
            }
        ]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个DescribeImagesResponse的实例，与请求对象对应
        resp = client.DescribeImages(req)
        # 输出json格式的字符串回包
        a = json.loads(resp.to_json_string())
        b = a['ImageSet']
        c = int(len(b) - 1)
        while c >= 0:
            print(b[c]['ImageId'], b[c]['ImageName'])
            c -= 1
        print()

    except TencentCloudSDKException as err:
        print(err)
def img_vps():
    Password = input("请输入密码，直接按enter默认Re123456789123") or "Re123456789123"
    print("-" * 5)
    print("idena镜像列表")
    img_namesearch()
    img = input("请输入镜像")
    print("-" * 5)
    print("自动今晚11：40注销服务器")
    print("请牢记密码为", Password)
    print("-" * 5)
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.ap-guangzhou.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.RunInstancesRequest()
        params = {
            "InstanceChargeType": "POSTPAID_BY_HOUR",
            "DisableApiTermination": False,
            "Placement": {
                "Zone": "ap-guangzhou-6",
                "ProjectId": 0
            },
            "VirtualPrivateCloud": {
                "AsVpcGateway": False,
                "VpcId": "vpc-b8ptxop5",
                "SubnetId": "subnet-ahm3l13o",
                "Ipv6AddressCount": 0
            },
            "InstanceType": "S6.MEDIUM8",
            "ImageId": img,
            "SystemDisk": {
                "DiskSize": 50,
                "DiskType": "CLOUD_BSSD"
            },
            "InternetAccessible": {
                "InternetMaxBandwidthOut": 100,
                "PublicIpAssigned": True,
                "InternetChargeType": "TRAFFIC_POSTPAID_BY_HOUR"
            },
            "LoginSettings": {
                "Password": Password
            },
            "SecurityGroupIds": ["sg-dbcg6j0n"],
            "InstanceCount": 1,
            "EnhancedService": {
                "SecurityService": {
                    "Enabled": False
                },
                "MonitorService": {
                    "Enabled": False
                },
                "AutomationService": {
                    "Enabled": True
                }
            },
            "ActionTimer": {
                "ActionTime": pro_base.today_day() + " 23:40:00",
                "TimerAction": "TerminateInstances",
                "Externals": {
                    "ReleaseAddress": True
                }
            },
            "TagSpecification": [
                {
                    "ResourceType": "instance",
                    "Tags": [
                        {
                            "Key": "idena",
                            "Value": "img"
                        }
                    ]
                }
            ]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个RunInstancesResponse的实例，与请求对象对应
        resp = client.RunInstances(req)
        # 输出json格式的字符串回包
        print(resp.to_json_string())

    except TencentCloudSDKException as err:
        print(err)
def img_build():
    img_vpsnamesearch()
    print("请输入vps名称")
    InstanceId=input()
    print("请输入镜像名称")
    ImageName=input()
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.CreateImageRequest()
        params = {
            "InstanceId": InstanceId,
            "ImageName": ImageName,

            "TagSpecification": [
                {
                    "ResourceType": "image",
                    "Tags": [
                        {
                            "Key": "idena",
                            "Value": "idena"
                        }
                    ]
                }
            ]

        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个CreateImageResponse的实例，与请求对象对应
        resp = client.CreateImage(req)
        # 输出json格式的字符串回包
        print(resp.to_json_string())

    except TencentCloudSDKException as err:
        print(err)

def vps_ip():
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.DescribeInstancesRequest()
        params = {
            "Limit": 100,
             "Filters": [
            {
                "Name": "tag-key",
                "Values": [ "idena" ]
            },
            {
                "Name": "tag-value",
                "Values": [ "idena" ]
            }
        ]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个DescribeInstancesResponse的实例，与请求对象对应
        resp = client.DescribeInstances(req)
        # 输出json格式的字符串回包

        a = json.loads(resp.to_json_string())
        b = a['InstanceSet']
        c = len(b) - 1

        with open(r'FWQ.prma', 'a+', encoding='utf-8') as test:
            test.truncate(0)

        print('所有服务器IP:')

        ip_addresses = []

        for i in range(c + 1):
            ip_addresses.append(b[i])

            write_to_excel(ip_addresses)
            write_to_file(ip_addresses)


        with open('FWQ.prma', 'r+') as f:
            content = f.read()
            f.seek(0, 0)
            f.write('[' + content)
        file = open("FWQ.prma", "a")
        file.write(']')
        file.close()

    except TencentCloudSDKException as err:
        print(err)

def write_to_file(ip_addresses):
    for ip in ip_addresses:
        json_str = f'''{{
            "UserName": "Administrator",
            "Password": "Re123456789123",
            "Address": "{ip['PublicIpAddresses'][0]}",
            "Port": "3389",
            "Protocol": "RDP",
            "ClassVersion": "RDP.V1",
            "DispName": "{ip['PublicIpAddresses'][0]}",
            "DisplayName": "{ip['PublicIpAddresses'][0]}",
        }},'''
        # 打开文件并将内容追加到末尾
    file = open("FWQ.prma", "a")
    file.write(json_str)
    file.close()


def write_to_excel(ip_addresses):
    workbook = openpyxl.load_workbook('idena-all.xlsx')
    sheet = workbook['Sheet1']

    for i, ip in enumerate(ip_addresses):
        ip_str = ','.join(ip['PublicIpAddresses'])
        print(ip_str)
        cell = sheet.cell(row=i + 2, column=3)  # Start from row 2 to leave the first row empty
        cell.value = ip_str

    workbook.save('idena-all.xlsx')



def idena_stats(i):
    def idena_balance(address):
        import requests
        import json
        url = 'https://api.idena.io/api/Address/' + '' + address
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url, headers=headers)
        a = json.loads(response.text)
        b = a['result']
        return b

    def idena_identity(address):
        import requests
        import json
        url = ' https://api.idena.io/api/Identity/' + '' + address
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url, headers=headers)
        a = json.loads(response.text)
        b = a['result']
        return b

    import openpyxl
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(state_excel)
    # 获取单元格数据
    sheet = workbook['Sheet1']
    cell = sheet['A' + '' + i]
    sheet.cell(row=int(i), column=4, value=idena_balance(cell.value)['balance'])
    sheet.cell(row=int(i), column=5, value=idena_balance(cell.value)['stake'])
    sheet.cell(row=int(i), column=6, value=idena_identity(cell.value)['state'])
    workbook.save(state_excel)

def ture_address():
    import openpyxl

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(state_excel)

    # 获取对应的工作表
    sheet = workbook.active

    # 遍历行数据，判断该行是否有有效数据
    row_count = 0
    for row in sheet.rows:
        if not all(cell.value is None for cell in row):
            row_count += 1
    return row_count
def img_ipsearch():
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = cvm_client.CvmClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.DescribeInstancesRequest()
        params = {
            "Filters": [
                {
                    "Name": "tag-key",
                    "Values": ["idena"]
                },
                {
                    "Name": "tag-value",
                    "Values": ["img"]
                }
            ]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个DescribeInstancesResponse的实例，与请求对象对应
        resp = client.DescribeInstances(req)

        a = json.loads(resp.to_json_string())
        b = a['InstanceSet']
        c = int(len(b) - 1)
        while c >= 0:
            print(b[c]['InstanceId'], b[c]['CreatedTime'], b[c]['PublicIpAddresses'])
            c -= 1
        print()
    except TencentCloudSDKException as err:
        print(err)
def img_nodestart():
    img=input("请输入服务器名称————ins开头")
    import json
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.tat.v20201028 import tat_client, models
    try:
        # 实例化一个认证对象，入参需要传入腾讯云账户 SecretId 和 SecretKey，此处还需注意密钥对的保密
        # 代码泄露可能会导致 SecretId 和 SecretKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考，建议采用更安全的方式来使用密钥，请参见：https://cloud.tencent.com/document/product/1278/85305
        # 密钥可前往官网控制台 https://console.cloud.tencent.com/cam/capi 进行获取
        cred = credential.Credential(secretId, secretKey)
        # 实例化一个http选项，可选的，没有特殊需求可以跳过
        httpProfile = HttpProfile()
        httpProfile.endpoint = "tat.tencentcloudapi.com"

        # 实例化一个client选项，可选的，没有特殊需求可以跳过
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        # 实例化要请求产品的client对象,clientProfile是可选的
        client = tat_client.TatClient(cred, "ap-guangzhou", clientProfile)

        # 实例化一个请求对象,每个接口都会对应一个request对象
        req = models.InvokeCommandRequest()
        params = {
            "CommandId": "cmd-p5wli3u3",
            "InstanceIds": [img]
        }
        req.from_json_string(json.dumps(params))

        # 返回的resp是一个InvokeCommandResponse的实例，与请求对象对应
        resp = client.InvokeCommand(req)
        # 输出json格式的字符串回包
        print(resp.to_json_string())

    except TencentCloudSDKException as err:
        print(err)
