import openpyxl
import paramiko
import time
idena_excel='idena-all.xlsx'
def main(i):
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(idena_excel)
    ws = wb.active

    # 获取 nodekey 和 hostname
    nodekey = ws.cell(row=i+1, column=2).value  # 第二行第二列
    hostname = ws.cell(row=i+1, column=3).value  # 第二行第三列

    # 服务器连接信息
    port = 22
    username = "administrator"
    password = "Re123456789123"

    sftp = None  # 初始化 sftp 变量

    try:
        # 连接服务器
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # 添加新的 SSH Key
        ssh.connect(hostname=hostname, port=port, username=username, password=password)

        # 修改文件内容
        filepath = r'C:\\Users\\Administrator\\Desktop\\datadir\\keystore\\nodekey'  # 文件路径
        new_content = nodekey.encode('utf-8')  # 要写入的内容
        sftp = ssh.open_sftp()
        with sftp.file(filepath, 'wb') as file:
            file.write(new_content)

        # 运行 nodekey.exe
        command = f"start C:\\Users\\Administrator\\Desktop\\nodekey.exe"
        stdin, stdout, stderr = ssh.exec_command(command)
        print(f"Executed command \"{command}\"")

    except paramiko.ssh_exception.AuthenticationException:
        print('Authentication failed.')
    except paramiko.ssh_exception.SSHException as e:
        print('SSH connection failed: ', e)
    except Exception as e:
        print('Error occurred while connecting to server or executing command: ', e)

    finally:
        # 关闭连接
        if sftp:
            sftp.close()
        ssh.close()

def ture_address():
    import openpyxl

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(idena_excel)

    # 获取对应的工作表
    sheet = workbook.active

    # 遍历行数据，判断该行是否有有效数据
    row_count = 0
    for row in sheet.rows:
        if not all(cell.value is None for cell in row):
            row_count += 1
    
    return row_count
print(ture_address())
i=0
while i<ture_address()-1:
    i += 1
    main(i)
