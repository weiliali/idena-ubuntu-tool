state_excel = 'idena-all.xlsx'
def idena_stats(i):
    def idena_balance(address):
        import requests
        import json
        url = 'https://api.idena.io/api/Address/' + '' + address
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url, headers=headers)
        a = json.loads(response.text)
        b = a['result']
        return b

    def idena_identity(address):
        import requests
        import json
        url = ' https://api.idena.io/api/Identity/' + '' + address
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url, headers=headers)
        a = json.loads(response.text)
        b = a['result']
        return b

    import openpyxl
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(state_excel)
    # 获取单元格数据
    sheet = workbook['Sheet1']
    cell = sheet['A' + '' + i]
    sheet.cell(row=int(i), column=4, value=idena_balance(cell.value)['balance'])
    sheet.cell(row=int(i), column=5, value=idena_balance(cell.value)['stake'])
    sheet.cell(row=int(i), column=6, value=idena_identity(cell.value)['state'])
    workbook.save(state_excel)


def ture_address():
    import openpyxl

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(state_excel)

    # 获取对应的工作表
    sheet = workbook.active

    # 遍历行数据，判断该行是否有有效数据
    row_count = 0
    for row in sheet.rows:
        if not all(cell.value is None for cell in row):
            row_count += 1
    return row_count


print("有效账户数", [ture_address() - 1])
m = 2
idena_stats(str(m))
while m < ture_address():
    m += 1
    idena_stats(str(m))
print("完成")
